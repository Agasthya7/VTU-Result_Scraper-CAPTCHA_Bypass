my_csv = open(r"Data\Subject_Codes.txt","r") #This file contains subjects and their credits    
my_text = my_csv.read()
sub_list = my_text.split(",")
sub_list = [sub_list.strip() for sub_list in sub_list]
#print(sub_list)
sub_code=[]# contains all the subject codes
sub_grade=[]# contains the credits for the above sujects

for i in range(0,len(sub_list)):
    if(i%2==0):
        sub_code.append(sub_list[i])
    else:
        sub_grade.append(int(sub_list[i]))
#print(sub_code)
#print(sub_grade)

#importing this at the top of the program causes issues with reading the text file
import openpyxl
from openpyxl import *
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import column_index_from_string, get_column_letter
from openpyxl.styles import numbers


xl_file=load_workbook(r"Data\Marks.xlsx")#opens the marks excel sheet
sheets=xl_file.sheetnames
#print(sheets)
#The below piece of code generates the grade point  for each subject and the no of credits and stores it in the corresponding rows in each worksheet
for sheet in sheets:
    worksheet=xl_file[sheet]
    subject_grade_index=sub_code.index(str(sheet))
    for row  in range(1,worksheet.max_row+1):
        marks=worksheet['E'+str(row)].value
        if marks >= 90 and marks <= 100:
            grade_point=10
        elif marks >=80 and marks < 90:
            grade_point=9
        elif marks >=70 and marks < 80:
            grade_point=8
        elif marks >=60 and marks < 70:
            grade_point=7
        elif marks >=50 and marks < 60:
            grade_point=6
        elif marks >=45 and marks < 50:
            grade_point=5
        elif marks >=40 and marks < 45:
            grade_point=4
        elif marks < 40 :
            grade_point=0
        #print(marks)
        worksheet['G'+str(row)]=grade_point*sub_grade[subject_grade_index]
        worksheet['H'+str(row)]=sub_grade[subject_grade_index]
xl_file.save(r"Data\Marks.xlsx")
'''The below piece of code handles creating a new worksheet named CGPA and storing the 
    usn and name of each student in that worksheet followed by adding their actual cgpa into the respective cell
'''
name_usn_dict={}#this dictionary holds the usn and name data extracted from other worksheets to put into cgpa worksheet
for sheet in sheets:
    worksheet=xl_file[sheet]
    for row in range (1,worksheet.max_row+1):
        usn=worksheet['A'+str(row)].value
        name=worksheet['B'+str(row)].value
        #if name_usn_dict[worksheet['A'+str(row)].value] not in name_usn_dict.keys():
        name_usn_dict[usn]=name
name_usn_dict=sorted(name_usn_dict.items())#Sorts the name and usn of the students into an order of the usn
#print(name_usn_dict)
xl_file.create_sheet('CGPA')

row=1
for key,value in name_usn_dict:
    worksheet=xl_file['CGPA']
    worksheet['A'+str(row)]=key
    worksheet['B'+str(row)]=value
    cixgi=0
    ci=0
    #the below for loop searches for name and usn in each worksheet and adds  it to compute cgpa
    for sheet in sheets:
        work_1=xl_file[sheet]
        for i in range(1,work_1.max_row+1):
            if work_1['A'+str(i)].value==key:
                #print('There is a match')
                cixgi=cixgi+int(work_1['G'+str(i)].value)
                ci=ci+int(work_1['H'+str(i)].value)
                break
    #print(cixgi,ci)           
    cgpa=round(cixgi/ci,2)   
    #print(key, cgpa)
    worksheet['C'+str(row)]=cgpa
    row=row+1
for sheet in sheets:
    worksheet=xl_file[sheet]
    worksheet.delete_cols(idx=7,amount=2) # this removes the gradepoint values and credits stored in all the previous worksheets to represent it neatly          
xl_file.save(r"Data\Marks.xlsx")