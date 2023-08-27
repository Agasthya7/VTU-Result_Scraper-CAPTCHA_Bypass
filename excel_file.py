#this code splits each line of the results.txt file into a element of the list
with open(r"Data\Results.txt", 'r') as file:
    lines = file.readlines()
data_list = [i for i in lines if i != '\n']#to remove spacing between two usn in text file
data_list = [data_list.strip() for data_list in data_list]

#the below code stores the sbject codes and their credits into their respective lists
my_csv = open(r"Data\Subject_Codes.txt","r")     
my_text = my_csv.read()
sub_list = my_text.split(",")
sub_list = [sub_list.strip() for sub_list in sub_list]
#print(sub_list)
sub_code=[]
sub_grade=[]

for i in range(0,len(sub_list)):
    if(i%2==0):
        sub_code.append(sub_list[i])
    else:
        sub_grade.append(int(sub_list[i]))
#print(sub_code)
#print(sub_grade)
import openpyxl
from openpyxl import *
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import column_index_from_string, get_column_letter
from openpyxl.styles import numbers


wb=load_workbook(r"Data\Marks.xlsx")
sheets=wb.sheetnames#Get the list of all the sheets present in the xl workbook
for sheet in sheets:
    wb.remove(wb[str(sheet)])# removes all the existing sheets in the workbook
for sub in sub_code:
    wb.create_sheet(str(sub))#creates a sheet for the suject in the sub_code list
    worksheet = wb[str(sub)]#loads the above created worksheet
    worksheet_row=1         #starts from the first  row to store the data
    for list_element in data_list: #goes through all the elemtns in the data_list which contains all the marks detail
        if sub in list_element:                                 #it checks if the selected subject is the one which is present in the list_element
            worksheet["A"+str(worksheet_row)] =list_element     #if yes it stores the data and then increases the row variable
            worksheet_row=worksheet_row+1
    '''Now all the data that is put into the excel sheets is present in the column 1 so now we split them 
        and store them in the consecutive columns. After storing them in the successive colums we remove the subject code and subject columns from them 
        as it is not necessary to have them and then save the workbook'''
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=1): 
        for cell in row:
            col_index = cell.col_idx
            col_letter = get_column_letter(col_index + 1)
            values = cell.value.split(",")
            for i, value in enumerate(values):
                cell = worksheet.cell(row=cell.row, column=col_index + i)
                if value.isnumeric():
                    cell.value = int(value)
                    cell.number_format = numbers.FORMAT_NUMBER
                else:
                    cell.value = value
    worksheet.delete_cols(3,amount=2)#deleting subject name and its code
    
wb.save(r"Data\Marks.xlsx")
