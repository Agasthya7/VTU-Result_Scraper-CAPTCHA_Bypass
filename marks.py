#This is the program that is written to calculate the total marks of the student.Do run the program after caclulating the cgpa using cgpa.py

import openpyxl
from openpyxl import *
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import column_index_from_string, get_column_letter
from openpyxl.styles import numbers


xl_file=load_workbook(r"Data\Marks.xlsx")#opens the marks excel sheet 
sheets=xl_file.sheetnames
sheets.remove('CGPA')
#print(sheets)

name_usn_dict={}#this dictionary holds the usn and name data extracted from other worksheets to put into cgpa worksheet
for sheet in sheets:
    worksheet=xl_file[sheet]
    for row in range (1,worksheet.max_row+1):
        usn=worksheet['A'+str(row)].value
        name=worksheet['B'+str(row)].value
        #if name_usn_dict[worksheet['A'+str(row)].value] not in name_usn_dict.keys():
        name_usn_dict[usn]=name
name_usn_dict=sorted(name_usn_dict.items())#Sorts the name and usn of the students into an order of the usn
#print(name_usn_dict)
xl_file.create_sheet('MARKS')


row=1
for key,value in name_usn_dict:
    worksheet=xl_file['MARKS']
    worksheet['A'+str(row)]=key
    worksheet['B'+str(row)]=value
    marks=0
    total=0
    #the below for loop searches for name and usn in each worksheet and adds  it to compute cgpa
    for sheet in sheets:
        work_1=xl_file[sheet]
        for i in range(1,work_1.max_row+1):
            if work_1['A'+str(i)].value==key:
                #print('There is a match')
                marks=marks+int(work_1['E'+str(i)].value)
                total=total+100
                break
    #print(marks,total)           
    percent=round((marks/total)*100,2)   
    #print(key, percent)
    worksheet['C'+str(row)]= marks#Use variable 'str(marks)+'/'+str(total)' to display marks in fraction
    worksheet['D'+str(row)]=percent
    row=row+1    
xl_file.save(r"Data\Marks.xlsx")